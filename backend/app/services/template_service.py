"""
Service Template - Génération de templates Excel pour import de données.

Sprint 2: Implémentation complète avec validation Excel intégrée.
"""
import logging
from datetime import datetime
from io import BytesIO
from typing import Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
    numbers
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.category import Category
from app.models.supplier import Supplier

logger = logging.getLogger(__name__)


class TemplateService:
    """
    Service pour générer des templates Excel personnalisés avec validation intégrée.

    Sprint 2: Version complète avec:
    - Validation Excel (listes déroulantes, formules)
    - Styling professionnel
    - Instructions détaillées
    - Exemples données
    """

    def __init__(self, db: Optional[Session] = None):
        """
        Initialiser le service template.

        Args:
            db: Session SQLAlchemy (optionnel pour récupérer catégories/fournisseurs)
        """
        self.db = db

    def generate_template(
        self,
        tenant_id: UUID,
        include_categories: bool = True,
        include_suppliers: bool = True,
        sample_data: bool = True,
    ) -> BytesIO:
        """
        Générer un template Excel pour import de données.

        Args:
            tenant_id: ID du tenant
            include_categories: Inclure onglet catégories
            include_suppliers: Inclure onglet fournisseurs
            sample_data: Inclure lignes d'exemple

        Returns:
            BytesIO contenant le fichier Excel
        """
        logger.info(f"Génération template pour tenant {tenant_id}")

        wb = Workbook()

        # Créer les onglets
        self._create_products_sheet(wb, tenant_id, sample_data)
        self._create_sales_sheet(wb, sample_data)
        self._create_instructions_sheet(wb)

        if include_categories:
            self._create_categories_sheet(wb, tenant_id)

        if include_suppliers:
            self._create_suppliers_sheet(wb, tenant_id)

        # Sauvegarder dans BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info("Template généré avec succès")
        return output

    def _create_products_sheet(self, wb: Workbook, tenant_id: UUID, sample_data: bool):
        """Créer l'onglet Produits avec validation."""
        ws = wb.active
        ws.title = "Produits"

        # En-têtes
        headers = [
            "Code*", "Nom*", "Catégorie*", "Fournisseur", "Prix Achat*",
            "Prix Vente*", "Unité*", "Stock Initial*", "Stock Min*", "Stock Max*",
            "Description", "Code-barres"
        ]

        # Style en-têtes
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = self._get_border()

        # Largeurs colonnes
        widths = [12, 25, 15, 15, 12, 12, 10, 13, 10, 10, 30, 15]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Hauteur ligne header
        ws.row_dimensions[1].height = 30

        # Validation liste déroulante pour Unité
        units = ["kg", "g", "L", "mL", "unité", "sac", "carton", "boîte", "paquet"]
        unit_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(units)}"',
            allow_blank=False,
            showDropDown=True
        )
        unit_validation.error = "Veuillez sélectionner une unité valide"
        unit_validation.errorTitle = "Unité invalide"
        ws.add_data_validation(unit_validation)
        unit_validation.add(f"G2:G1000")  # Colonne Unité

        # Validation numérique pour prix et stocks
        for col in ["E", "F", "H", "I", "J"]:  # Prix et stocks
            num_validation = DataValidation(
                type="decimal",
                operator="greaterThan",
                formula1=0,
                allow_blank=False
            )
            num_validation.error = "La valeur doit être un nombre positif"
            num_validation.errorTitle = "Valeur invalide"
            ws.add_data_validation(num_validation)
            num_validation.add(f"{col}2:{col}1000")

        # Données d'exemple
        if sample_data:
            examples = [
                ["RIZ001", "Riz 50kg Sénégal", "Céréales", "SONACOS", 25000, 27500, "sac", 50, 10, 100, "Riz brisé 50kg importé", ""],
                ["HUILE002", "Huile 20L", "Condiments", "SUNEOR", 18000, 20000, "L", 30, 5, 80, "Huile d'arachide 20L", ""],
                ["SUC003", "Sucre 1kg", "Épicerie", "CSS", 650, 750, "kg", 200, 50, 300, "Sucre cristallisé 1kg", ""],
            ]

            for row_idx, example in enumerate(examples, start=2):
                for col_idx, value in enumerate(example, start=1):
                    cell = ws.cell(row_idx, col_idx, value)
                    cell.border = self._get_border(style="thin")
                    cell.alignment = Alignment(vertical="center")

                    # Format nombres
                    if col_idx in [5, 6]:  # Prix
                        cell.number_format = '#,##0'
                    elif col_idx in [8, 9, 10]:  # Stocks
                        cell.number_format = '0'

        # Figer la première ligne
        ws.freeze_panes = "A2"

        # Commentaire sur header
        ws["A1"].comment = None  # Supprimer commentaire existant si présent

    def _create_sales_sheet(self, wb: Workbook, sample_data: bool):
        """Créer l'onglet Ventes avec validation."""
        ws = wb.create_sheet("Ventes")

        # En-têtes
        headers = ["Code Produit*", "Date Vente*", "Quantité*", "Prix Unitaire*", "Montant Total"]

        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = self._get_border()

        # Largeurs colonnes
        widths = [15, 15, 12, 15, 15]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[1].height = 30

        # Validation date
        date_validation = DataValidation(
            type="date",
            operator="lessThan",
            formula1=datetime.now().strftime("%Y-%m-%d"),
            allow_blank=False
        )
        date_validation.error = "La date doit être au format YYYY-MM-DD et antérieure à aujourd'hui"
        date_validation.errorTitle = "Date invalide"
        ws.add_data_validation(date_validation)
        date_validation.add("B2:B10000")

        # Validation quantité et prix
        for col in ["C", "D"]:
            num_validation = DataValidation(
                type="decimal",
                operator="greaterThan",
                formula1=0,
                allow_blank=False
            )
            num_validation.error = "La valeur doit être un nombre positif"
            num_validation.errorTitle = "Valeur invalide"
            ws.add_data_validation(num_validation)
            num_validation.add(f"{col}2:{col}10000")

        # Formule pour Montant Total
        if sample_data:
            examples = [
                ["RIZ001", "2025-10-01", 5, 27500, "=C2*D2"],
                ["HUILE002", "2025-10-03", 10, 20000, "=C3*D3"],
                ["SUC003", "2025-10-05", 20, 750, "=C4*D4"],
            ]

            for row_idx, example in enumerate(examples, start=2):
                for col_idx, value in enumerate(example, start=1):
                    cell = ws.cell(row_idx, col_idx, value)
                    cell.border = self._get_border(style="thin")
                    cell.alignment = Alignment(vertical="center")

                    if col_idx == 2:  # Date
                        cell.number_format = 'yyyy-mm-dd'
                    elif col_idx in [3]:  # Quantité
                        cell.number_format = '0'
                    elif col_idx in [4, 5]:  # Prix et montant
                        cell.number_format = '#,##0'

        ws.freeze_panes = "A2"

    def _create_instructions_sheet(self, wb: Workbook):
        """Créer l'onglet Instructions."""
        ws = wb.create_sheet("Instructions")

        # Titre
        ws["A1"] = "📋 GUIDE D'UTILISATION DU TEMPLATE D'IMPORT DIGIBOOST PME"
        ws["A1"].font = Font(bold=True, size=14, color="4F46E5")
        ws.merge_cells("A1:E1")
        ws.row_dimensions[1].height = 25

        instructions = [
            "",
            "📌 ÉTAPES D'UTILISATION",
            "",
            "1️⃣  Remplissez l'onglet 'Produits' avec votre catalogue produits",
            "    → Les colonnes marquées * sont OBLIGATOIRES",
            "    → Le code produit doit être UNIQUE",
            "",
            "2️⃣  Remplissez l'onglet 'Ventes' avec vos ventes historiques (optionnel)",
            "    → Utilisez les codes produits saisis dans l'onglet Produits",
            "    → Historique recommandé: minimum 30 jours de ventes",
            "",
            "3️⃣  Vérifiez vos données:",
            "    → Pas de doublons dans les codes produits",
            "    → Tous les prix sont positifs",
            "    → Les dates sont au format YYYY-MM-DD",
            "",
            "4️⃣  Importez ce fichier dans l'étape 4 du wizard d'onboarding",
            "",
            "",
            "📊 FORMAT DES DONNÉES",
            "",
            "ONGLET PRODUITS:",
            "  • Code*: Identifiant unique (ex: RIZ001, HUILE002)",
            "  • Nom*: Nom du produit (max 255 caractères)",
            "  • Catégorie*: Catégorie du produit (sera créée si inexistante)",
            "  • Fournisseur: Nom du fournisseur (optionnel)",
            "  • Prix Achat*: Prix d'achat en FCFA (nombre sans séparateur)",
            "  • Prix Vente*: Prix de vente en FCFA",
            "  • Unité*: Unité de mesure (liste déroulante fournie)",
            "  • Stock Initial*: Quantité en stock actuel",
            "  • Stock Min*: Seuil d'alerte rupture",
            "  • Stock Max*: Stock maximum recommandé",
            "",
            "ONGLET VENTES:",
            "  • Code Produit*: Doit correspondre à un code dans l'onglet Produits",
            "  • Date Vente*: Format YYYY-MM-DD (ex: 2025-10-23)",
            "  • Quantité*: Nombre entier positif",
            "  • Prix Unitaire*: Prix de vente unitaire en FCFA",
            "  • Montant Total: Calculé automatiquement (Quantité × Prix)",
            "",
            "",
            "⚠️ ERREURS COURANTES À ÉVITER",
            "",
            "❌ Code produit dupliqué → Chaque code doit être unique",
            "❌ Prix négatif ou zéro → Tous les prix doivent être > 0",
            "❌ Stock max < stock min → Stock max doit être supérieur au stock min",
            "❌ Date future → Les dates de vente doivent être passées",
            "❌ Code produit inexistant dans Ventes → Vérifier onglet Produits",
            "",
            "",
            "💡 CONSEILS",
            "",
            "✅ Commencez avec 10-20 produits pour tester",
            "✅ Utilisez les données d'exemple fournies comme modèle",
            "✅ Sauvegardez régulièrement votre fichier",
            "✅ Vérifiez les validations Excel (listes déroulantes, formats)",
            "",
            "",
            "📞 SUPPORT",
            "",
            "En cas de problème lors de l'import:",
            "→ Vérifiez le rapport d'erreurs généré",
            "→ Contactez support@digiboost.sn",
            "",
            f"Template généré le: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        ]

        for row, instruction in enumerate(instructions, start=2):
            ws.cell(row, 1, instruction)
            if instruction.startswith("📌") or instruction.startswith("📊") or instruction.startswith("⚠️") or instruction.startswith("💡") or instruction.startswith("📞"):
                ws.cell(row, 1).font = Font(bold=True, size=12, color="1F2937")
            elif instruction.startswith("  •"):
                ws.cell(row, 1).font = Font(size=10)

        ws.column_dimensions["A"].width = 100

    def _create_categories_sheet(self, wb: Workbook, tenant_id: UUID):
        """Créer l'onglet Catégories (référence)."""
        ws = wb.create_sheet("Catégories (Référence)")

        ws["A1"] = "Catégorie"
        ws["A1"].font = Font(bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")

        # Catégories par défaut
        default_categories = [
            "Céréales", "Condiments", "Épicerie", "Boissons",
            "Produits laitiers", "Viandes", "Fruits & Légumes",
            "Hygiène", "Entretien", "Divers"
        ]

        # Récupérer catégories existantes du tenant si DB disponible
        if self.db:
            try:
                existing = self.db.query(Category).filter(
                    Category.tenant_id == tenant_id
                ).all()
                categories = [c.name for c in existing] if existing else default_categories
            except Exception:
                categories = default_categories
        else:
            categories = default_categories

        for idx, cat in enumerate(categories, start=2):
            ws.cell(idx, 1, cat)

        ws.column_dimensions["A"].width = 20

    def _create_suppliers_sheet(self, wb: Workbook, tenant_id: UUID):
        """Créer l'onglet Fournisseurs (référence)."""
        ws = wb.create_sheet("Fournisseurs (Référence)")

        ws["A1"] = "Fournisseur"
        ws["A1"].font = Font(bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")

        # Fournisseurs par défaut
        default_suppliers = ["SONACOS", "SUNEOR", "CSS", "Mimran", "Autres"]

        # Récupérer fournisseurs existants du tenant si DB disponible
        if self.db:
            try:
                existing = self.db.query(Supplier).filter(
                    Supplier.tenant_id == tenant_id
                ).all()
                suppliers = [s.name for s in existing] if existing else default_suppliers
            except Exception:
                suppliers = default_suppliers
        else:
            suppliers = default_suppliers

        for idx, sup in enumerate(suppliers, start=2):
            ws.cell(idx, 1, sup)

        ws.column_dimensions["A"].width = 20

    def _get_border(self, style: str = "medium") -> Border:
        """Créer un style de bordure."""
        side_style = Side(style=style, color="D1D5DB")
        return Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
