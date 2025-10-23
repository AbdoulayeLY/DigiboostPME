"""
Service Import - Validation et import de données Excel.

Sprint 2: Validation complète de fichiers Excel avec Pandas.
"""
import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from uuid import UUID

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.category import Category
from app.models.product import Product
from app.models.supplier import Supplier

logger = logging.getLogger(__name__)

# Codes d'erreur standardisés
ERROR_FILE_FORMAT = "ERR_FILE_001"
ERROR_STRUCTURE = "ERR_STRUCT_002"
ERROR_VALIDATION = "ERR_VALID_003"
ERROR_IMPORT = "ERR_IMPORT_004"


class ImportService:
    """
    Service pour valider et importer des données depuis Excel.

    Étapes:
    1. validate_excel_structure(): Vérifier onglets et headers
    2. validate_products(): Vérifier données produits
    3. validate_sales(): Vérifier données ventes
    4. generate_validation_report(): Générer rapport d'erreurs
    """

    def __init__(self, db: Session):
        """Initialiser le service import."""
        self.db = db

    def validate_excel_file(
        self, file_path: str, tenant_id: UUID
    ) -> Tuple[bool, Dict]:
        """
        Valider complètement un fichier Excel.

        Args:
            file_path: Chemin vers le fichier Excel
            tenant_id: ID du tenant

        Returns:
            Tuple (is_valid, report)
        """
        errors = []
        warnings = []
        stats = {}

        try:
            # 1. Vérifier structure
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames

            if "Produits" not in sheet_names:
                errors.append({
                    "code": ERROR_STRUCTURE,
                    "message": "Onglet 'Produits' manquant",
                    "sheet": None,
                })
                return False, self._generate_report(False, errors, warnings, stats)

            # 2. Charger DataFrames
            df_products = pd.read_excel(file_path, sheet_name="Produits")
            df_sales = None
            if "Ventes" in sheet_names:
                df_sales = pd.read_excel(file_path, sheet_name="Ventes")

            # 3. Valider headers produits
            required_headers = [
                "Code*", "Nom*", "Catégorie*", "Prix Achat*",
                "Prix Vente*", "Unité*", "Stock Initial*", "Stock Min*", "Stock Max*"
            ]

            # Headers flexibles (accepter avec ou sans *)
            actual_headers = df_products.columns.tolist()
            normalized_headers = [h.replace("*", "").strip() for h in actual_headers]
            required_normalized = [h.replace("*", "").strip() for h in required_headers]

            missing = [h for h in required_normalized if h not in normalized_headers]
            if missing:
                errors.append({
                    "code": ERROR_STRUCTURE,
                    "message": f"Colonnes manquantes: {', '.join(missing)}",
                    "sheet": "Produits",
                })
                return False, self._generate_report(False, errors, warnings, stats)

            # 4. Valider produits
            products_errors, products_warnings = self._validate_products(df_products, tenant_id)
            errors.extend(products_errors)
            warnings.extend(products_warnings)

            # 5. Valider ventes si présentes
            if df_sales is not None and not df_sales.empty:
                sales_errors, sales_warnings = self._validate_sales(df_sales, df_products)
                errors.extend(sales_errors)
                warnings.extend(sales_warnings)
                stats["sales_count"] = len(df_sales)
            else:
                stats["sales_count"] = 0

            stats["products_count"] = len(df_products)

            is_valid = len(errors) == 0
            return is_valid, self._generate_report(is_valid, errors, warnings, stats)

        except Exception as e:
            logger.error(f"Erreur validation Excel: {str(e)}")
            errors.append({
                "code": ERROR_FILE_FORMAT,
                "message": f"Erreur lecture fichier: {str(e)}",
                "sheet": None,
            })
            return False, self._generate_report(False, errors, warnings, stats)

    def _validate_products(
        self, df: pd.DataFrame, tenant_id: UUID
    ) -> Tuple[List[Dict], List[Dict]]:
        """Valider données produits."""
        errors = []
        warnings = []

        # Normaliser noms colonnes
        df.columns = [col.replace("*", "").strip() for col in df.columns]

        # Vérifier doublons codes
        codes = df["Code"].dropna()
        duplicates = codes[codes.duplicated()].unique()
        if len(duplicates) > 0:
            for code in duplicates:
                rows = df[df["Code"] == code].index.tolist()
                errors.append({
                    "code": ERROR_VALIDATION,
                    "sheet": "Produits",
                    "rows": [r + 2 for r in rows],  # +2 car header = row 1
                    "column": "Code",
                    "message": f"Code produit dupliqué: {code}",
                    "value": code,
                })

        # Valider chaque ligne
        for idx, row in df.iterrows():
            row_num = idx + 2

            # Code obligatoire
            if pd.isna(row.get("Code")):
                errors.append({
                    "code": ERROR_VALIDATION,
                    "sheet": "Produits",
                    "row": row_num,
                    "column": "Code",
                    "message": "Code produit obligatoire",
                })

            # Nom obligatoire
            if pd.isna(row.get("Nom")):
                errors.append({
                    "code": ERROR_VALIDATION,
                    "sheet": "Produits",
                    "row": row_num,
                    "column": "Nom",
                    "message": "Nom produit obligatoire",
                })

            # Prix > 0
            for col in ["Prix Achat", "Prix Vente"]:
                val = row.get(col)
                if pd.isna(val) or val <= 0:
                    errors.append({
                        "code": ERROR_VALIDATION,
                        "sheet": "Produits",
                        "row": row_num,
                        "column": col,
                        "message": f"{col} doit être > 0",
                        "value": val,
                    })

            # Stocks >= 0
            for col in ["Stock Initial", "Stock Min", "Stock Max"]:
                val = row.get(col)
                if pd.isna(val) or val < 0:
                    errors.append({
                        "code": ERROR_VALIDATION,
                        "sheet": "Produits",
                        "row": row_num,
                        "column": col,
                        "message": f"{col} doit être >= 0",
                        "value": val,
                    })

            # Stock Max > Stock Min
            stock_min = row.get("Stock Min", 0)
            stock_max = row.get("Stock Max", 0)
            if not pd.isna(stock_min) and not pd.isna(stock_max):
                if stock_max < stock_min:
                    errors.append({
                        "code": ERROR_VALIDATION,
                        "sheet": "Produits",
                        "row": row_num,
                        "column": "Stock Max",
                        "message": "Stock Max doit être >= Stock Min",
                        "value": f"max={stock_max}, min={stock_min}",
                    })

            # Warning stock élevé
            if not pd.isna(row.get("Stock Initial")) and row["Stock Initial"] > 1000:
                warnings.append({
                    "sheet": "Produits",
                    "row": row_num,
                    "column": "Stock Initial",
                    "message": "Stock initial élevé (>1000)",
                    "value": row["Stock Initial"],
                })

        return errors, warnings

    def _validate_sales(
        self, df_sales: pd.DataFrame, df_products: pd.DataFrame
    ) -> Tuple[List[Dict], List[Dict]]:
        """Valider données ventes."""
        errors = []
        warnings = []

        # Normaliser colonnes
        df_sales.columns = [col.replace("*", "").strip() for col in df_sales.columns]
        product_codes = set(df_products["Code"].dropna().str.strip())

        for idx, row in df_sales.iterrows():
            row_num = idx + 2

            # Code produit existe
            code = row.get("Code Produit")
            if pd.isna(code):
                errors.append({
                    "code": ERROR_VALIDATION,
                    "sheet": "Ventes",
                    "row": row_num,
                    "column": "Code Produit",
                    "message": "Code produit obligatoire",
                })
            elif str(code).strip() not in product_codes:
                errors.append({
                    "code": ERROR_VALIDATION,
                    "sheet": "Ventes",
                    "row": row_num,
                    "column": "Code Produit",
                    "message": f"Code produit inexistant: {code}",
                    "value": code,
                })

            # Date valide
            try:
                date_vente = pd.to_datetime(row.get("Date Vente"))
                if date_vente > datetime.now():
                    errors.append({
                        "code": ERROR_VALIDATION,
                        "sheet": "Ventes",
                        "row": row_num,
                        "column": "Date Vente",
                        "message": "Date vente future",
                        "value": str(date_vente),
                    })
            except Exception:
                errors.append({
                    "code": ERROR_VALIDATION,
                    "sheet": "Ventes",
                    "row": row_num,
                    "column": "Date Vente",
                    "message": "Format date invalide (attendu: YYYY-MM-DD)",
                    "value": row.get("Date Vente"),
                })

            # Quantité > 0
            qty = row.get("Quantité")
            if pd.isna(qty) or qty <= 0:
                errors.append({
                    "code": ERROR_VALIDATION,
                    "sheet": "Ventes",
                    "row": row_num,
                    "column": "Quantité",
                    "message": "Quantité doit être > 0",
                    "value": qty,
                })

            # Prix unitaire > 0
            price = row.get("Prix Unitaire")
            if pd.isna(price) or price <= 0:
                errors.append({
                    "code": ERROR_VALIDATION,
                    "sheet": "Ventes",
                    "row": row_num,
                    "column": "Prix Unitaire",
                    "message": "Prix unitaire doit être > 0",
                    "value": price,
                })

        return errors, warnings

    def _generate_report(
        self, is_valid: bool, errors: List[Dict], warnings: List[Dict], stats: Dict
    ) -> Dict:
        """Générer rapport de validation."""
        return {
            "valid": is_valid,
            "errors": errors,
            "warnings": warnings,
            "stats": stats,
            "error_count": len(errors),
            "warning_count": len(warnings),
        }
