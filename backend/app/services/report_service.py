"""
Service pour la génération de rapports (Excel et PDF).
"""
from sqlalchemy.orm import Session
from sqlalchemy import text, func
from uuid import UUID
from typing import Dict, Any, Optional
from datetime import datetime, timedelta
from io import BytesIO
import calendar

# Excel imports
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart

# PDF imports
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

# Matplotlib imports
import matplotlib
matplotlib.use('Agg')  # Backend non-interactif pour serveur
import matplotlib.pyplot as plt

# Models
from app.models.product import Product
from app.models.sale import Sale
from app.models.category import Category


class ReportService:
    """Service pour générer les rapports automatisés."""

    def __init__(self, db: Session):
        self.db = db

    def generate_inventory_report(self, tenant_id: UUID) -> BytesIO:
        """
        Rapport Inventaire Stock (Excel).

        Colonnes:
        - Code
        - Nom
        - Catégorie
        - Stock Actuel
        - Stock Min
        - Stock Max
        - Statut
        - Valorisation

        Args:
            tenant_id: UUID du tenant

        Returns:
            BytesIO: Fichier Excel en mémoire
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventaire Stock"

        # En-tête rapport
        ws['A1'] = "RAPPORT INVENTAIRE STOCK"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Date: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        # En-têtes colonnes
        headers = [
            "Code", "Nom Produit", "Catégorie", "Stock Actuel",
            "Stock Min", "Stock Max", "Unité", "Statut",
            "Prix Achat", "Valorisation"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Données
        products = self.db.query(Product).filter(
            Product.tenant_id == tenant_id,
            Product.is_active == True
        ).order_by(Product.name).all()

        row = 5
        for product in products:
            # Calculer statut
            status = self._calculate_status(product)
            valorisation = product.current_stock * product.purchase_price

            ws.cell(row, 1, product.code)
            ws.cell(row, 2, product.name)
            ws.cell(row, 3, product.category.name if product.category else "-")
            ws.cell(row, 4, float(product.current_stock))
            ws.cell(row, 5, float(product.min_stock or 0))
            ws.cell(row, 6, float(product.max_stock or 0))
            ws.cell(row, 7, product.unit)
            ws.cell(row, 8, status)
            ws.cell(row, 9, float(product.purchase_price))
            ws.cell(row, 10, float(valorisation))

            # Coloration statut
            status_cell = ws.cell(row, 8)
            if status == "RUPTURE":
                status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                status_cell.font = Font(color="991B1B", bold=True)
            elif status == "FAIBLE":
                status_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                status_cell.font = Font(color="92400E", bold=True)
            elif status == "ALERTE":
                status_cell.fill = PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid")
                status_cell.font = Font(color="9A3412", bold=True)

            row += 1

        # Totaux
        row += 1
        ws.cell(row, 9, "TOTAL:")
        ws.cell(row, 9).font = Font(bold=True)
        ws.cell(row, 10, f"=SUM(J5:J{row-2})")
        ws.cell(row, 10).font = Font(bold=True)
        ws.cell(row, 10).number_format = '#,##0'

        # Ajuster largeurs colonnes
        column_widths = [12, 30, 20, 12, 12, 12, 10, 15, 15, 18]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row_cells in ws[f'A4:J{row}']:
            for cell in row_cells:
                cell.border = thin_border

        # Sauvegarder dans BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def generate_sales_analysis_report(
        self,
        tenant_id: UUID,
        start_date: datetime,
        end_date: datetime
    ) -> BytesIO:
        """
        Rapport Analyse Ventes (Excel multi-onglets).

        Onglets:
        1. Synthèse
        2. Ventes par Produit
        3. Ventes par Catégorie
        4. Évolution Quotidienne

        Args:
            tenant_id: UUID du tenant
            start_date: Date de début
            end_date: Date de fin

        Returns:
            BytesIO: Fichier Excel en mémoire
        """
        wb = openpyxl.Workbook()

        # ONGLET 1: Synthèse
        ws_summary = wb.active
        ws_summary.title = "Synthèse"

        # Période
        ws_summary['A1'] = "ANALYSE VENTES"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A2'] = f"Période: {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}"

        # KPIs
        kpis = self.db.query(
            func.count(Sale.id).label('transactions'),
            func.sum(Sale.quantity).label('units'),
            func.sum(Sale.total_amount).label('revenue')
        ).filter(
            Sale.tenant_id == tenant_id,
            Sale.sale_date >= start_date,
            Sale.sale_date <= end_date
        ).first()

        ws_summary['A4'] = "Indicateurs Clés"
        ws_summary['A4'].font = Font(bold=True, size=12)

        ws_summary['A6'] = "Nombre de transactions:"
        ws_summary['B6'] = kpis.transactions or 0
        ws_summary['B6'].font = Font(bold=True)

        ws_summary['A7'] = "Unités vendues:"
        ws_summary['B7'] = float(kpis.units or 0)
        ws_summary['B7'].font = Font(bold=True)

        ws_summary['A8'] = "Chiffre d'affaires:"
        ws_summary['B8'] = float(kpis.revenue or 0)
        ws_summary['B8'].font = Font(bold=True)
        ws_summary['B8'].number_format = '#,##0 "FCFA"'

        ws_summary['A9'] = "Panier moyen:"
        if kpis.transactions and kpis.transactions > 0:
            ws_summary['B9'] = float(kpis.revenue or 0) / kpis.transactions
        else:
            ws_summary['B9'] = 0
        ws_summary['B9'].font = Font(bold=True)
        ws_summary['B9'].number_format = '#,##0 "FCFA"'

        # ONGLET 2: Ventes par Produit
        ws_products = wb.create_sheet("Ventes par Produit")

        headers = ["Code", "Nom Produit", "Catégorie", "Quantité", "CA", "Transactions"]
        for col, header in enumerate(headers, start=1):
            cell = ws_products.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Données ventes par produit
        query = text("""
            SELECT
                p.code,
                p.name,
                c.name as category,
                SUM(s.quantity) as quantity,
                SUM(s.total_amount) as revenue,
                COUNT(s.id) as transactions
            FROM products p
            LEFT JOIN sales s ON p.id = s.product_id
                AND s.sale_date >= :start_date
                AND s.sale_date <= :end_date
            LEFT JOIN categories c ON p.category_id = c.id
            WHERE p.tenant_id = :tenant_id
                AND s.id IS NOT NULL
            GROUP BY p.id, p.code, p.name, c.name
            ORDER BY SUM(s.total_amount) DESC
        """)

        results = self.db.execute(query, {
            "tenant_id": str(tenant_id),
            "start_date": start_date,
            "end_date": end_date
        }).fetchall()

        for row_idx, result in enumerate(results, start=2):
            ws_products.cell(row_idx, 1, result.code)
            ws_products.cell(row_idx, 2, result.name)
            ws_products.cell(row_idx, 3, result.category or "-")
            ws_products.cell(row_idx, 4, float(result.quantity))
            ws_products.cell(row_idx, 5, float(result.revenue))
            ws_products.cell(row_idx, 5).number_format = '#,##0'
            ws_products.cell(row_idx, 6, result.transactions)

        # Ajuster largeurs colonnes
        ws_products.column_dimensions['A'].width = 12
        ws_products.column_dimensions['B'].width = 30
        ws_products.column_dimensions['C'].width = 20
        ws_products.column_dimensions['D'].width = 12
        ws_products.column_dimensions['E'].width = 18
        ws_products.column_dimensions['F'].width = 15

        # Graphique Top 10
        if len(results) > 0:
            chart = BarChart()
            chart.title = "Top 10 Produits (CA)"
            chart.x_axis.title = "Produits"
            chart.y_axis.title = "CA (FCFA)"

            data = Reference(ws_products, min_col=5, min_row=1, max_row=min(11, len(results)+1))
            cats = Reference(ws_products, min_col=2, min_row=2, max_row=min(11, len(results)+1))

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws_products.add_chart(chart, "H2")

        # ONGLET 3: Ventes par Catégorie
        ws_categories = wb.create_sheet("Ventes par Catégorie")

        headers_cat = ["Catégorie", "Produits", "Quantité", "CA", "Transactions"]
        for col, header in enumerate(headers_cat, start=1):
            cell = ws_categories.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        query_cat = text("""
            SELECT
                COALESCE(c.name, 'Sans catégorie') as category,
                COUNT(DISTINCT p.id) as product_count,
                SUM(s.quantity) as quantity,
                SUM(s.total_amount) as revenue,
                COUNT(s.id) as transactions
            FROM sales s
            JOIN products p ON s.product_id = p.id
            LEFT JOIN categories c ON p.category_id = c.id
            WHERE s.tenant_id = :tenant_id
                AND s.sale_date >= :start_date
                AND s.sale_date <= :end_date
            GROUP BY c.name
            ORDER BY SUM(s.total_amount) DESC
        """)

        results_cat = self.db.execute(query_cat, {
            "tenant_id": str(tenant_id),
            "start_date": start_date,
            "end_date": end_date
        }).fetchall()

        for row_idx, result in enumerate(results_cat, start=2):
            ws_categories.cell(row_idx, 1, result.category)
            ws_categories.cell(row_idx, 2, result.product_count)
            ws_categories.cell(row_idx, 3, float(result.quantity))
            ws_categories.cell(row_idx, 4, float(result.revenue))
            ws_categories.cell(row_idx, 4).number_format = '#,##0'
            ws_categories.cell(row_idx, 5, result.transactions)

        # Ajuster largeurs
        ws_categories.column_dimensions['A'].width = 25
        ws_categories.column_dimensions['B'].width = 12
        ws_categories.column_dimensions['C'].width = 12
        ws_categories.column_dimensions['D'].width = 18
        ws_categories.column_dimensions['E'].width = 15

        # ONGLET 4: Évolution Quotidienne
        ws_daily = wb.create_sheet("Évolution Quotidienne")

        headers_daily = ["Date", "Transactions", "CA"]
        for col, header in enumerate(headers_daily, start=1):
            cell = ws_daily.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        query_daily = text("""
            SELECT
                DATE(sale_date) as date,
                COUNT(id) as transactions,
                SUM(total_amount) as revenue
            FROM sales
            WHERE tenant_id = :tenant_id
                AND sale_date >= :start_date
                AND sale_date <= :end_date
            GROUP BY DATE(sale_date)
            ORDER BY DATE(sale_date)
        """)

        results_daily = self.db.execute(query_daily, {
            "tenant_id": str(tenant_id),
            "start_date": start_date,
            "end_date": end_date
        }).fetchall()

        for row_idx, result in enumerate(results_daily, start=2):
            ws_daily.cell(row_idx, 1, result.date.strftime('%d/%m/%Y'))
            ws_daily.cell(row_idx, 2, result.transactions)
            ws_daily.cell(row_idx, 3, float(result.revenue))
            ws_daily.cell(row_idx, 3).number_format = '#,##0'

        # Ajuster largeurs
        ws_daily.column_dimensions['A'].width = 15
        ws_daily.column_dimensions['B'].width = 15
        ws_daily.column_dimensions['C'].width = 18

        # Graphique évolution
        if len(results_daily) > 0:
            chart = LineChart()
            chart.title = "Évolution du CA"
            chart.x_axis.title = "Date"
            chart.y_axis.title = "CA (FCFA)"

            data = Reference(ws_daily, min_col=3, min_row=1, max_row=len(results_daily)+1)
            cats = Reference(ws_daily, min_col=1, min_row=2, max_row=len(results_daily)+1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws_daily.add_chart(chart, "E2")

        # Sauvegarder
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def _calculate_status(self, product: Product) -> str:
        """
        Calculer statut stock produit.

        Args:
            product: Produit à analyser

        Returns:
            str: Statut (RUPTURE, FAIBLE, ALERTE, SURSTOCK, NORMAL)
        """
        if product.current_stock == 0:
            return "RUPTURE"
        elif product.min_stock and product.current_stock <= product.min_stock:
            return "FAIBLE"
        elif product.min_stock and product.current_stock <= float(product.min_stock) * 1.2:
            return "ALERTE"
        elif product.max_stock and product.current_stock >= product.max_stock:
            return "SURSTOCK"
        return "NORMAL"

    def generate_monthly_summary_pdf(
        self,
        tenant_id: UUID,
        month: int,
        year: int
    ) -> BytesIO:
        """
        Rapport Synthèse Mensuelle (PDF).

        Contenu:
        - En-tête avec période
        - KPIs principaux (grid)
        - Graphique évolution CA
        - Top 5 produits
        - Alertes stock
        - Recommandations

        Args:
            tenant_id: UUID du tenant
            month: Mois (1-12)
            year: Année

        Returns:
            BytesIO: Fichier PDF en mémoire
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4F46E5'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1F2937'),
            spaceAfter=12,
            spaceBefore=20
        )

        # Éléments du document
        story = []

        # Titre
        month_names = [
            "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
            "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
        ]
        title = Paragraph(
            f"SYNTHÈSE MENSUELLE<br/>{month_names[month-1]} {year}",
            title_style
        )
        story.append(title)
        story.append(Spacer(1, 1*cm))

        # Date génération
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_RIGHT
        )
        story.append(Paragraph(
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
            date_style
        ))
        story.append(Spacer(1, 0.5*cm))

        # Calculer dates début/fin du mois
        start_date = datetime(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = datetime(year, month, last_day, 23, 59, 59)

        # KPIs principaux
        story.append(Paragraph("Indicateurs Clés", heading_style))

        # Calculer KPIs ventes
        kpis = self.db.query(
            func.count(Sale.id).label('transactions'),
            func.sum(Sale.total_amount).label('revenue')
        ).filter(
            Sale.tenant_id == tenant_id,
            Sale.sale_date >= start_date,
            Sale.sale_date <= end_date
        ).first()

        # Santé stock
        stock_health = self.db.query(
            func.count(Product.id).label('total')
        ).filter(
            Product.tenant_id == tenant_id,
            Product.is_active == True
        ).first()

        rupture_count = self.db.query(
            func.count(Product.id)
        ).filter(
            Product.tenant_id == tenant_id,
            Product.is_active == True,
            Product.current_stock == 0
        ).scalar()

        faible_count = self.db.query(
            func.count(Product.id)
        ).filter(
            Product.tenant_id == tenant_id,
            Product.is_active == True,
            Product.current_stock > 0,
            Product.current_stock <= Product.min_stock
        ).scalar()

        # Table KPIs
        kpi_data = [
            ['Indicateur', 'Valeur'],
            ['Chiffre d\'Affaires', f'{int(kpis.revenue or 0):,} FCFA'.replace(',', ' ')],
            ['Transactions', f'{kpis.transactions or 0}'],
            ['Panier Moyen', f'{int((kpis.revenue or 0) / max(kpis.transactions or 1, 1)):,} FCFA'.replace(',', ' ')],
            ['', ''],
            ['Produits Actifs', f'{stock_health.total or 0}'],
            ['Ruptures Stock', f'{rupture_count or 0}'],
            ['Stock Faible', f'{faible_count or 0}'],
        ]

        kpi_table = Table(kpi_data, colWidths=[8*cm, 8*cm])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        story.append(kpi_table)
        story.append(Spacer(1, 1*cm))

        # Graphique évolution CA
        story.append(Paragraph("Évolution du Chiffre d'Affaires", heading_style))

        # Requête données quotidiennes
        query = text("""
            SELECT
                DATE(sale_date) as date,
                SUM(total_amount) as revenue
            FROM sales
            WHERE tenant_id = :tenant_id
                AND sale_date >= :start_date
                AND sale_date <= :end_date
            GROUP BY DATE(sale_date)
            ORDER BY date
        """)

        daily_sales = self.db.execute(query, {
            "tenant_id": str(tenant_id),
            "start_date": start_date,
            "end_date": end_date
        }).fetchall()

        if daily_sales and len(daily_sales) > 0:
            # Créer graphique matplotlib
            fig, ax = plt.subplots(figsize=(12, 6))
            dates = [row.date for row in daily_sales]
            revenues = [float(row.revenue) for row in daily_sales]

            ax.plot(dates, revenues, marker='o', linewidth=2, color='#4F46E5', markersize=6)
            ax.set_xlabel('Date', fontsize=12)
            ax.set_ylabel('CA (FCFA)', fontsize=12)
            ax.grid(True, alpha=0.3)
            ax.ticklabel_format(style='plain', axis='y')

            # Formater axe Y
            from matplotlib.ticker import FuncFormatter
            def format_func(value, tick_number):
                if value >= 1000000:
                    return f'{int(value/1000000)}M'
                elif value >= 1000:
                    return f'{int(value/1000)}K'
                return f'{int(value)}'

            ax.yaxis.set_major_formatter(FuncFormatter(format_func))

            # Rotation labels dates
            plt.xticks(rotation=45, ha='right')

            # Sauvegarder dans buffer
            img_buffer = BytesIO()
            plt.tight_layout()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            img_buffer.seek(0)
            plt.close()

            # Ajouter au PDF
            img = Image(img_buffer, width=16*cm, height=8*cm)
            story.append(img)
        else:
            story.append(Paragraph("Aucune donnée de ventes pour cette période", styles['Normal']))

        story.append(Spacer(1, 1*cm))

        # Top 5 produits
        story.append(Paragraph("Top 5 Produits du Mois", heading_style))

        top_query = text("""
            SELECT
                p.name,
                SUM(s.quantity) as quantity,
                SUM(s.total_amount) as revenue
            FROM products p
            JOIN sales s ON p.id = s.product_id
            WHERE p.tenant_id = :tenant_id
                AND s.sale_date >= :start_date
                AND s.sale_date <= :end_date
            GROUP BY p.id, p.name
            ORDER BY SUM(s.total_amount) DESC
            LIMIT 5
        """)

        top_products = self.db.execute(top_query, {
            "tenant_id": str(tenant_id),
            "start_date": start_date,
            "end_date": end_date
        }).fetchall()

        if top_products and len(top_products) > 0:
            top_data = [['Produit', 'Quantité', 'CA (FCFA)']]
            for product in top_products:
                top_data.append([
                    product.name,
                    f'{float(product.quantity):.1f}',
                    f'{int(product.revenue):,}'.replace(',', ' ')
                ])

            top_table = Table(top_data, colWidths=[8*cm, 4*cm, 4*cm])
            top_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            story.append(top_table)
        else:
            story.append(Paragraph("Aucune vente enregistrée pour cette période", styles['Normal']))

        story.append(Spacer(1, 1*cm))

        # Alertes Stock
        story.append(Paragraph("Alertes Stock", heading_style))

        # Produits en rupture ou stock faible
        alert_products = self.db.query(Product).filter(
            Product.tenant_id == tenant_id,
            Product.is_active == True,
            Product.current_stock <= Product.min_stock
        ).order_by(Product.current_stock).limit(10).all()

        if alert_products:
            alert_data = [['Produit', 'Stock Actuel', 'Stock Min', 'Statut']]
            for product in alert_products:
                status = "RUPTURE" if product.current_stock == 0 else "FAIBLE"
                alert_data.append([
                    product.name,
                    f'{float(product.current_stock):.1f}',
                    f'{float(product.min_stock or 0):.1f}',
                    status
                ])

            alert_table = Table(alert_data, colWidths=[6*cm, 3*cm, 3*cm, 4*cm])
            alert_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EF4444')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            story.append(alert_table)
        else:
            story.append(Paragraph("✓ Aucune alerte stock ce mois", styles['Normal']))

        # Footer personnalisé
        def footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 9)
            canvas.setFillColor(colors.grey)
            canvas.drawString(2*cm, 1*cm, "DigiboostPME - Intelligence Supply Chain")
            canvas.drawRightString(A4[0] - 2*cm, 1*cm, f"Page {doc.page}")
            canvas.restoreState()

        # Générer PDF
        try:
            doc.build(story, onFirstPage=footer, onLaterPages=footer)
        finally:
            self.db.rollback()

        buffer.seek(0)
        return buffer
